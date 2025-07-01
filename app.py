import streamlit as st
import pdfplumber
import pandas as pd
import io
import json
from google import genai
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import requests
import base64
import os
from PIL import Image
import fitz 




def upload_to_imagekit(image_path):
    # ده هو الـ endpoint الصحيح للرفع
    url = "https://upload.imagekit.io/api/v1/files/upload"

    with open(image_path, "rb") as f:
        files = {"file": f}
        data = {
            "fileName": os.path.basename(image_path),
            "folder": "/cv_images", 
        }

        # إعداد الـ Authorization header باستخدام base64 من private key
        private_key = st.secrets["IMAGEKIT_API"]
        encoded = base64.b64encode(f"{private_key}:".encode()).decode()
        headers = {"Authorization": f"Basic {encoded}"}

        res = requests.post(url, files=files, data=data, headers=headers)

        if res.status_code == 200:
            return res.json()["url"]  # ده اللينك المباشر للصورة
        else:
            print("❌ Error uploading image:", res.status_code, res.text)
            return None

client = genai.Client(api_key=st.secrets["GEMINI_API_KEY"])

# System prompt for Gemini
system_prompt = """
You are a resume parser. Extract structured data from any CV text and return valid JSON that maps to the following schema.

Include only fields that can be extracted directly from the CV. Omit any system-generated fields like IDs.

Map links like LinkedIn, GitHub, and Portfolio to their correct fields. If the text includes a label (e.g., “LinkedIn:”) followed by a non-clickable name, but embedded links are listed below under [Resolved Links], use the actual URLs.

Return all dates in YYYY-MM-DD format, and normalize phone numbers to international format (e.g., +[CountryCode]-[Number]).

Return ONLY valid JSON. Do not include explanation or markdown. Start with '{' and end with '}'.

it is important to add each field in it's corrent section, for example if the CV has a Arabic but not in language it should be add to the language section 


ERD:
 
	1.	Candidate
	•	CandidateID (Primary Key)
	•	FullName
	•	Nationality
	•	CurrentLocation
	•	Phone
	•	Email
	•	LinkedInURL
	•	CareerSummary
	•	ProfilePhoto (Base64 encoded string)
	•	PortfolioLink
	2.	EmploymentHistory
	•	EmploymentID (Primary Key)
	•	CandidateID (Foreign Key)
	•	JobTitle
	•	Company
	•	Location
	•	StartDate
	•	EndDate (note add present if still employed)
	•	Responsibilities
	3.	Education
	•	EducationID (Primary Key)
	•	CandidateID (Foreign Key)
	•	Degree (e.g., Bachelor's, Master's)
	•	Institution
	•	Location (e.g., City, Country , or check your  database for this name or search for it and add hint not included in CV)
	•	GraduationDate
	•	Major   (e.g. , Computer Science, Business Administration)
    •	ProjectName (max 2 to 5 words) (e.g., Final Year Project , or try to extract from the CV or conclusion)
    •	ProjectDescription (e.g., A web application for managing student projects)
	4.	Certifications
	•	CertificationID (Primary Key)
	•	CandidateID (Foreign Key)
	•	CertificationTitle
	•	IssuingOrganization
	•	IssueDate
	•	ExpiryDate
	5.	Skills
	•	SkillID (Primary Key)
	•	CandidateID (Foreign Key)
	•	SkillName
	•	ProficiencyLevel (e.g., Beginner, Intermediate, Advanced)
	6.	Projects
	•	ProjectID (Primary Key)
	•	CandidateID (Foreign Key)
	•	ProjectTitle
	•	ProjectDescription
	•	Role
	•	Duration
	•	TechnologiesUsed
	7.	Publications
	•	PublicationID (Primary Key)
	•	CandidateID (Foreign Key)
	•	PublicationTitle
	•	PublicationDate
	•	Publisher
	•	Description
	8.	VolunteerExperience
	•	VolunteerID (Primary Key)
	•	CandidateID (Foreign Key)
	•	Organization
	•	Role
	•	Duration
	•	ActivitiesImpact
	9.	References
	•	ReferenceID (Primary Key)
	•	CandidateID (Foreign Key)
	•	ReferenceName
	•	Position
	•	ContactInformation
	•	RelationToCandidate
	10.	OtherInformation
 
	•	OtherInfoID (Primary Key)
	•	CandidateID (Foreign Key)
	•	InformationType (e.g., hobbies, languages, portfolio link)
	•	Details
 
	11.	Languages
 
	•	LanguageID (Primary Key)
	•	CandidateID (Foreign Key)
	•	LanguageName
	•	ProficiencyLevel (e.g., Native, Fluent, Intermediate, Beginner)
 
	12.	Awards
 
	•	AwardID (Primary Key)
	•	CandidateID (Foreign Key)
	•	AwardTitle
	•	IssuingOrganization
	•	AwardDate
	•	Description
 
	13.	Interests
 
	•	InterestID (Primary Key)
	•	CandidateID (Foreign Key)
	•	InterestName
	•	Description
"""

st.title("📄 CV Parser - Kafaat solution")
uploaded_files = st.file_uploader("Upload one or more CVs (PDF)", type="pdf", accept_multiple_files=True)

if uploaded_files and st.button("Parse CVs"):
    excel_buffer = io.BytesIO()
    writer = pd.ExcelWriter(excel_buffer, engine="openpyxl")
    IMAGE_DIR = "extracted_images"
    os.makedirs(IMAGE_DIR, exist_ok=True)
    preview_rows = []
    with st.spinner(f"🔄 Processing {len(uploaded_files)} CVs... Please wait"):
        for idx, file in enumerate(uploaded_files, 1):
            st.write(f"Processing {idx} of {len(uploaded_files)}: {file.name}")
        # for file in uploaded_files:
            # image extraction
            # ay7aga
            profile_image_path = ""
       
           

            file_bytes = file.read()
            doc = fitz.open(stream=io.BytesIO(file_bytes), filetype="pdf")

            # Extract best (largest) image
            max_area = 0
            for page_index in range(len(doc)):
                images = doc.get_page_images(page_index)
                for img_index, img in enumerate(images):
                    xref = img[0]
                    width, height = img[2], img[3]
                    area = width * height

                    base_image = doc.extract_image(xref)
                    image_bytes = base_image["image"]
                    image_ext = base_image["ext"]

                    image_filename = f"profile_page{page_index+1}_img{img_index+1}.{image_ext}"
                    image_path = os.path.join(IMAGE_DIR, image_filename)

                    with open(image_path, "wb") as f:
                        f.write(image_bytes)

                    if area > max_area:
                        max_area = area
                        profile_image_path = image_path

            # TEXT EXTRACTION
            # with pdfplumber.open(file) as pdf:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:

                text_lines = []
                if profile_image_path:

                    image_url = upload_to_imagekit(profile_image_path) 
                    text_lines.append(f"Embedded Link: ProfilePhoto  : {image_url}")
                for page in pdf.pages:
                    text_lines.append(page.extract_text() or "")
                    for link in page.hyperlinks:
                        uri = link.get("uri", "")
                        if uri:
                            text_lines.append(f"Embedded Link: {uri}")
            text = "\n".join(text_lines)

            response = client.models.generate_content(
                model="gemini-1.5-flash",
                config={"response_mime_type": "application/json"},
                contents=system_prompt + "\n\nCV Content:\n" + text,
            )

            try:
                parsed = json.loads(response.text.strip())
                candidate_name = parsed.get("Candidate", {}).get("FullName", "Unknown").strip().replace(" ", "_")

                # Build table-like structure
                flat_rows = []
                for section, content in parsed.items():
                    if isinstance(content, dict):
                        for k, v in content.items():
                            flat_rows.append({"Section": section, "Field": k, "Value": v})
                    elif isinstance(content, list):
                        for item in content:
                            if isinstance(item, dict):
                                for k, v in item.items():
                                    flat_rows.append({"Section": section, "Field": k, "Value": v})
                                flat_rows.append({"Section": "", "Field": "", "Value": ""})  # spacer
                            else:
                                flat_rows.append({"Section": section, "Field": "", "Value": item})
                    else:
                        flat_rows.append({"Section": section, "Field": "", "Value": content})

                df_flat = pd.DataFrame(flat_rows)


                df_flat.to_excel(writer, sheet_name=candidate_name[:31], index=False)

                # احصل على الـ worksheet قبل الإغلاق
                ws = writer.book[candidate_name[:31]]

                # التنسيقات
                fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                bold = Font(bold=True)

                section = None
                for row in range(2, ws.max_row + 1):
                    cell = ws[f"A{row}"]
                    if cell.value and cell.value != section:
                        section = cell.value
                        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                        ws[f"A{row}"].fill = fill
                        ws[f"A{row}"].font = bold

                # الآن احفظ التعديلات
                writer.close()
                excel_buffer.seek(0)
                # wb.save(excel_buffer)
                # excel_buffer.seek(0)

                preview_rows.append({
                    "File": file.name,
                    "Candidate": parsed.get("Candidate", {}).get("FullName", "")
                })

            except Exception as e:
                st.error(f"❌ Failed to parse {file.name}: {e}")
                st.code(response.text)

    # Preview summary
    if preview_rows:
        st.dataframe(pd.DataFrame(preview_rows))

    writer.close()
    excel_buffer.seek(0)

    st.download_button(
        label="📥 Download Organized Excel",
        data=excel_buffer,
        file_name="organized_cv_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
